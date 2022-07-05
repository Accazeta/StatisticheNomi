import traceback
from openpyxl import load_workbook

import functions as f

try:
    wb = load_workbook('Nomi_Neonati_Bergamo.xlsx')
    ws = wb.active

    print("\nQueste sono le info disponibili")
    try:
        for row in ws.iter_rows(min_col=ws.min_column, max_col=ws.max_column, min_row=ws.min_row, max_row=ws.min_row):
            for cell in row:
                if cell.value == "Nascita_anno".upper():
                    print('|ANNO', end='')
                else:
                    print('|{}'.format(cell.value), end='')
        print('|')
    except:
        print(traceback.format_exc())

    try:
        # nome = input("Inserisci il nome che vuoi cercare: ")
        lista_anni = f.check_available_years(ws)
        gap_years = f.check_missing_years(lista_anni)
        if gap_years is None:
            print("Analizzo i dati dal {} al {}".format(lista_anni[0], lista_anni[-1]))
        else:
            print("Analizzo i dati dal {} al {}, tuttavia mancano questi anni: {}".format(lista_anni[0], lista_anni[-1],
                                                                                          gap_years))

        scelta = 999
        while scelta != 0:
            print(
                "\nMenu: \n1) Ricerca per nome\n2) Ricerca per anno\n3) Ricerca per nazionalità\n4) Ricerca per "
                "genere\n0) per chiudere")
            scelta = int(input("Scelta: "))
            match scelta:
                case 1:
                    nome = input("Inserisci il nome da cercare: ")
                    f.check_by_name(ws, nome=nome)
                # case 2:
                # codice
                # case 3:
                # codice
                # case 4:
                # codice
                # case 0:
                # codice
                case _:
                    scelta = 999
    except:
        print(traceback.format_exc())
except FileNotFoundError:
    print("Errore, file non trovato")
except:
    print(traceback.format_exc())
